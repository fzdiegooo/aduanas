import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

MAX_WORKERS = 4

BASE_URL  = "http://www.aduanet.gob.pe/cl-ad-itconsmanifiesto/manifiestoITS01Alias"
BASE_HOST = "http://www.aduanet.gob.pe"

COLUMNAS_EXCEL = [
    "Manifiesto", "Fecha de Salida", "Aerolínea", "Vuelo",
    "Puerto de Embarque", "BL", "Fecha de Transmisión",
    "Bultos", "Peso Bruto", "Empaques", "Embarcador", "Consignatario",
    "Marcas y Números", "Descripción de Mercadería",
]

# ============================================================
#  CLASIFICACIÓN  (integrado de Producto_aereo.py)
# ============================================================

PALABRAS_CLAVE_PRODUCTOS = {
    "Jengibre fresco": ["jengibre", "ginger", "organic ginger",
                        "fresh organic ginger", "organic fresh ginger"],
    "Cúrcuma":   ["turmeric", "curcuma", "cúrcuma", "organic turmeric"],
    "Palta":     ["palta", "paltas", "avocado", "aguacate", "aguacates"],
    "Arándano":  ["arándano", "arándanos", "blueberries"],
    "Naranja":   ["naranja", "naranjas", "orange", "oranges"],
    "Café":      ["cafe", "café", "cafes", "coffee"],
    "Cacao":     ["cacao", "cocoa"],
    "Uva":       ["uva", "uvas", "grape", "grapes"],
    "Mango":     ["mango", "mangos"],
    "Achiote":   ["achiote", "annatto"],
    "Ajo":       ["ajo", "garlic"],
    "Limón":     ["limon", "lemon", "limón", "lemons"],
    "Esparrago": ["esparrago", "esparragos"],
    "Fresas":    ["fresa", "strawberries", "fresas"],
    "Quinua":    ["quinua", "quinoa"],
    "Mandarina": ["mandarina", "clementinas", "clementine"],
    "Menestras": ["beans", "lentejas", "menestras", "frejoles", "alberja"],
    "Cebolla":   ["cebolla", "onion", "onions", "cebollas"],
}

EXPORTADORES_ESPECIFICOS = [
    "vancard", "elisur", "fruitxchange", "hamillton", "la grama", "fruticola selva",
    "namaskar", "masojo", "anawi", "jungle fresh organic", "la campiña",
    "jch exportaciones", "blue pacific oils", "aseptic peruvian fruit",
]

IMPORTADORES_PAISES = {
    "AXARFRUIT":                           "ESPAÑA",
    "JALHUCA EXPLOTACIONES SL":            "ESPAÑA",
    "PIMENTON Y DERIVADOS S.L CALLE CATA": "ESPAÑA",
    "CASCONE":                             "PAISES BAJOS",
    "FRUCHTHANSA":                         "PAISES BAJOS",
    "IPOKI BV":                            "PAISES BAJOS",
    "NATURE S PRODUCE SP Z":               "PAISES BAJOS",
    "NFG New Fruit Group":                 "PAISES BAJOS",
    "VISION INTERNATIONAL B.V.":           "PAISES BAJOS",
    "AGROFAIR":                            "ITALIA",
    "ANAWI USA":                           "ESTADOS UNIDOS",
    "DELINA":                              "ESTADOS UNIDOS",
    "ECORIPE TROPICALS":                   "ESTADOS UNIDOS",
    "GLOBAL FARMS ENTERPRISES":            "ESTADOS UNIDOS",
    "HEATH AND LEJEUNE":                   "ESTADOS UNIDOS",
    "I LOVE PRODUCE":                      "ESTADOS UNIDOS",
    "INTERNATIONAL SPECIALTY PRODUCE":     "ESTADOS UNIDOS",
    "IPOKI PRODUCE LLC":                   "ESTADOS UNIDOS",
    "J Y C TROPICALS INC":                 "ESTADOS UNIDOS",
    "JLZ PRODUCE":                         "ESTADOS UNIDOS",
    "SUNDINE PRODUCE INC":                 "ESTADOS UNIDOS",
    "TRINITY DISTRIBUTION INC":            "ESTADOS UNIDOS",
    "UREN NORTH AMERICA LLC":              "ESTADOS UNIDOS",
    "VIVA TIERRA ORGANIC INC":             "ESTADOS UNIDOS",
    "THOMAS FRESH INC":                    "CANADÁ",
    "SOL FRUIT IMPORTS LTD":               "REINO UNIDO",
}

CONTINENTES = {
    "ESPAÑA":         "Europa",
    "PAISES BAJOS":   "Europa",
    "ITALIA":         "Europa",
    "ESTADOS UNIDOS": "América",
    "CANADÁ":         "América",
    "REINO UNIDO":    "Europa",
}


def _clean(text):
    return text.replace("\xa0", " ").strip()


def _identificar_producto(descripcion, embarcador, consignatario):
    desc = descripcion.lower() if isinstance(descripcion, str) else ""
    for producto, palabras in PALABRAS_CLAVE_PRODUCTOS.items():
        if any(p.lower() in desc for p in palabras):
            return producto
    if isinstance(embarcador, str):
        for exp in EXPORTADORES_ESPECIFICOS:
            if exp in embarcador.lower():
                return "¿Jengibre?"
    if isinstance(consignatario, str):
        for imp in IMPORTADORES_PAISES:
            if imp.lower() in consignatario.lower():
                return "¿Jengibre?"
    return "No asociado"


def _segmentar_jengibre(descripcion, producto):
    if producto == "Jengibre fresco" and isinstance(descripcion, str):
        d = descripcion.lower()
        if "juice" in d or "jugo" in d:
            return "Jugo de jengibre"
        if "bags" in d or "sacos" in d or "bolsas" in d:
            return "Jengibre deshidratado"
    return producto


def _identificar_pais(consignatario):
    if not isinstance(consignatario, str):
        return ""
    for imp, pais in IMPORTADORES_PAISES.items():
        if imp.lower() in consignatario.lower():
            return pais
    return ""


# ============================================================
#  FORMATO EXCEL
# ============================================================

def _aplicar_formato_excel(ruta, df, nombre_tabla):
    wb = load_workbook(ruta)
    ws = wb.active
    n_cols = len(df.columns)
    n_rows = len(df)
    col_fin = get_column_letter(n_cols)
    ref = f"A1:{col_fin}{n_rows + 1}"

    tabla = Table(displayName=nombre_tabla, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=True,
    )
    ws.add_table(tabla)
    ws.auto_filter.ref = ref

    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        ws.row_dimensions[row[0].row].height = 15
        for cell in row:
            cell.alignment = Alignment(wrap_text=False)

    wb.save(ruta)


# ============================================================
#  CLASIFICACIÓN
# ============================================================

def _clasificar_df(df, directorio_base, nombre_archivo_base):
    nombre_completo = nombre_archivo_base.replace("detalle_aereo_", "detalle_aereo_completo_", 1)
    ruta = os.path.join(directorio_base, nombre_completo + ".xlsx")

    fechas = pd.to_datetime(df["Fecha de Salida"], dayfirst=True, errors="coerce")
    df.insert(1, "Envio",  "Aereo")
    df.insert(2, "Semana", fechas.dt.isocalendar().week.astype("Int64"))
    df.insert(3, "Año",    fechas.dt.year.astype("Int64"))

    df["Producto"] = df.apply(
        lambda r: _segmentar_jengibre(
            r.get("Descripción de Mercadería", ""),
            _identificar_producto(
                r.get("Descripción de Mercadería", ""),
                r.get("Embarcador", ""),
                r.get("Consignatario", ""),
            ),
        ),
        axis=1,
    )
    df["TIPO"]          = ""
    df["Pais"]          = df["Consignatario"].apply(_identificar_pais)
    df["Ciudad destino"] = ""
    df["Continente"]    = df["Pais"].apply(lambda p: CONTINENTES.get(p, "") if p else "")

    df.to_excel(ruta, index=False)
    _aplicar_formato_excel(ruta, df, "DetalleAereoCompleto")
    print(f"Archivo clasificado: {ruta}")


# ============================================================
#  NIVEL 3 — detalle de mercadería (llamado desde ThreadPoolExecutor)
# ============================================================

def _fetch_nivel3(session, headers,
                  val_manifiesto, val_fecha_salida, val_aerolinea,
                  val_vuelo, val_puerto, val_bl, val_fecha_transmision,
                  anno, numero_manif, numdet, numcon, numconm, tipomani):

    payload = {
        "accion":       "consultarDetalleConocimientoEmbarqueExportacion",
        "CG_cadu":      "235",
        "CMc2_Anno":    anno,
        "CMc2_Numero":  numero_manif,
        "CMc2_numcon":  numcon,
        "CMc2_numconm": numconm,
        "CMc2_NumDet":  numdet,
        "CMc2_TipM":    tipomani,
        "tipo_archivo": "",
        "reporte":      "ExpAerGuia",
        "backPage":     "ConsulManifExpAerFechList",
    }
    try:
        resp = session.post(BASE_URL, data=payload, headers=headers, timeout=30)
        resp.encoding = "ISO-8859-1"
        soup = BeautifulSoup(resp.text, "html.parser")
    except Exception as exc:
        print(f"    Error nivel 3 BL={val_bl}: {exc}")
        return []

    filas = []
    for tr in soup.find_all("tr", class_="bg"):
        tds = tr.find_all("td", recursive=False)
        if len(tds) != 7:
            continue

        # Saltear fila vacía inicial
        bultos = _clean(tds[0].get_text(strip=True))
        if not bultos:
            continue

        peso_bruto    = _clean(tds[1].get_text(strip=True))
        empaques      = _clean(tds[2].get_text(strip=True))
        embarcador    = _clean(tds[3].get_text(strip=True))
        if "Ley 29733" in embarcador:
            embarcador = ""
        consignatario = _clean(tds[4].get_text(strip=True))
        marcas        = _clean(tds[5].get_text(" ", strip=True))
        descripcion   = _clean(tds[6].get_text(strip=True))

        filas.append([
            val_manifiesto, val_fecha_salida, val_aerolinea, val_vuelo, val_puerto,
            val_bl, val_fecha_transmision,
            bultos, peso_bruto, empaques, embarcador, consignatario,
            marcas, descripcion,
        ])

    return filas


# ============================================================
#  PROCESO PRINCIPAL
# ============================================================

def procesar_manifiestos(fecha_inicio, fecha_fin):
    session = requests.Session()
    headers = {
        "User-Agent":      "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
        "Content-Type":    "application/x-www-form-urlencoded",
        "Referer":         BASE_URL,
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "es-PE,es;q=0.9,en;q=0.8",
    }

    directorio_base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    os.makedirs(directorio_base, exist_ok=True)

    partes = fecha_inicio.split("/")
    if fecha_inicio == fecha_fin:
        nombre_archivo = f"detalle_aereo_{partes[0]}-{partes[1]}-{partes[2]}"
    else:
        fi = fecha_inicio.replace("/", "-")
        ff = fecha_fin.replace("/", "-")
        nombre_archivo = f"detalle_aereo_{fi}_al_{ff}"

    ruta_raw = os.path.join(directorio_base, nombre_archivo + ".xlsx")
    t0 = time.time()

    # ── Inicializar sesión (cookies JSESSIONID) ──────────────────────────────
    session.get(
        BASE_URL,
        params={"accion": "cargaConsultaManifiesto", "tipoConsulta": "fechaSalida"},
        headers=headers, timeout=30,
    )

    # ── NIVEL 1: lista de manifiestos (con paginación) ───────────────────────
    print("Consultando manifiestos...")
    manifiestos   = []
    seen_mani     = set()
    pagina_n1     = 1
    MAX_PAGES_N1  = 200

    while pagina_n1 <= MAX_PAGES_N1:
        if pagina_n1 == 1:
            resp1 = session.post(
                BASE_URL,
                data={
                    "accion":       "consultaManifiesto",
                    "fec_inicio":   fecha_inicio,
                    "fec_fin":      fecha_fin,
                    "cod_terminal": "0000",
                },
                headers=headers, timeout=30,
            )
        else:
            resp1 = session.post(
                f"{BASE_HOST}/cl-ad-itconsmanifiesto/ConsulManifExpAerFechList.jsp",
                data={"tamanioPagina": "10", "pagina": str(pagina_n1 - 1)},
                headers=headers, timeout=30,
            )
        resp1.encoding = "ISO-8859-1"
        soup1 = BeautifulSoup(resp1.text, "html.parser")

        nuevos = 0
        for link in soup1.find_all("a", href=re.compile(r"jsDetalle2")):
            m = re.search(r"jsDetalle2\('(\d+)','(\d+)'\)", link.get("href", ""))
            if not m:
                continue
            anio, numero = m.groups()
            clave = (anio, numero)
            if clave in seen_mani:
                continue
            seen_mani.add(clave)
            tr = link.find_parent("tr")
            if not tr:
                continue
            tds = tr.find_all("td")
            if len(tds) < 6:
                continue
            manifiestos.append({
                "texto":        _clean(link.get_text(strip=True)),
                "anio":         anio,
                "numero":       numero,
                "fecha_salida": _clean(tds[1].get_text(strip=True)),
                "aerolinea":    _clean(tds[3].get_text(strip=True)),
                "puerto":       _clean(tds[4].get_text(strip=True)),
                "vuelo":        _clean(tds[5].get_text(strip=True)),
            })
            nuevos += 1

        if nuevos == 0:
            break

        mc1 = re.search(r"(\d+)\s+a\s+(\d+)\s+de\s+(\d+)", soup1.get_text())
        if mc1 and int(mc1.group(2)) < int(mc1.group(3)):
            pagina_n1 += 1
        else:
            break

    print(f"Manifiestos encontrados: {len(manifiestos)}")

    df_final    = pd.DataFrame(columns=COLUMNAS_EXCEL)
    total_filas = 0

    for idx, mani in enumerate(manifiestos, 1):
        print(f"[{idx}/{len(manifiestos)}] Manifiesto {mani['texto']}")
        anno_full = str(2000 + int(mani["anio"]))   # "26" → "2026"

        # ── NIVEL 2: guías del manifiesto ────────────────────────────────────
        guias      = []
        pagina_det = 1
        seen_bls   = set()
        MAX_PAGES  = 100
        while pagina_det <= MAX_PAGES:
            if pagina_det == 1:
                resp2 = session.post(
                    BASE_URL,
                    data={
                        "accion":        "consultaManifiestoGuia",
                        "CMc1_Anno":     f"00{mani['anio']}",   # "0026"
                        "CMc1_Numero":   mani["numero"],
                        "CMc1_Terminal": "0000",
                        "viat":          "4",
                        "CG_cadu":       "235",
                    },
                    headers=headers, timeout=30,
                )
            else:
                resp2 = session.post(
                    f"{BASE_HOST}/cl-ad-itconsmanifiesto/ConsulManifExpAerGuia.jsp",
                    data={"tamanioPagina": "10", "pagina": str(pagina_det - 1)},
                    headers=headers, timeout=30,
                )
            resp2.encoding = "ISO-8859-1"
            soup2 = BeautifulSoup(resp2.text, "html.parser")

            guias_pag = []
            for link_d in soup2.find_all("a", href=re.compile(r"jsDetalleD")):
                href = link_d.get("href", "")
                md = re.search(
                    r"jsDetalleD\('([^']+)','([^']+)','([^']*)','([^']+)'\)", href
                )
                if not md:
                    continue
                numdet, numcon, tipomani, numconm = md.groups()

                tr2 = link_d.find_parent("tr")
                if not tr2:
                    continue
                tds2 = tr2.find_all("td")
                if len(tds2) < 13:
                    continue

                bl_a = tds2[0].find("a")
                bl   = _clean(bl_a.get_text(strip=True)) if bl_a else numcon.strip()
                ft   = _clean(tds2[12].get_text(strip=True))

                # Clave única por guía+detalle para detectar paginación duplicada
                clave = (bl.strip(), numdet.strip())
                if clave in seen_bls:
                    continue
                seen_bls.add(clave)

                guias_pag.append({
                    "bl":                bl,
                    "numdet":            numdet,
                    "numcon":            numcon,
                    "numconm":           numconm,
                    "tipomani":          tipomani,
                    "fecha_transmision": ft,
                })

            # Si esta página no aportó guías nuevas → fin de paginación
            if not guias_pag:
                break

            guias.extend(guias_pag)

            # Paginación nivel 2
            mc = re.search(r"(\d+)\s+a\s+(\d+)\s+de\s+(\d+)", soup2.get_text())
            if mc and int(mc.group(2)) < int(mc.group(3)):
                pagina_det += 1
            else:
                break

        if not guias:
            print("  Sin guías")
            continue
        print(f"  Guías: {len(guias)}")

        # ── NIVEL 3: detalles en paralelo ────────────────────────────────────
        filas_mani = []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as exe:
            futs = {
                exe.submit(
                    _fetch_nivel3, session, headers,
                    mani["texto"], mani["fecha_salida"], mani["aerolinea"],
                    mani["vuelo"], mani["puerto"],
                    g["bl"], g["fecha_transmision"],
                    anno_full, mani["numero"],
                    g["numdet"], g["numcon"], g["numconm"], g["tipomani"],
                ): g for g in guias
            }
            for fut in as_completed(futs):
                res = fut.result()
                if res:
                    filas_mani.extend(res)

        if filas_mani:
            df_final = pd.concat(
                [df_final, pd.DataFrame(filas_mani, columns=COLUMNAS_EXCEL)],
                ignore_index=True,
            )
            total_filas += len(filas_mani)
            # Guardado incremental
            df_final.to_excel(ruta_raw, index=False)
            print(f"  +{len(filas_mani)} → total {total_filas} | {ruta_raw}")

    if df_final.empty:
        print("Sin datos.")
        return

    # Guardado final raw con formato
    df_final.sort_values(
        by=["Manifiesto", "Fecha de Salida"], inplace=True, ignore_index=True
    )
    df_final.to_excel(ruta_raw, index=False)
    _aplicar_formato_excel(ruta_raw, df_final, "DetalleAereoRaw")
    print(f"\nArchivo raw: {ruta_raw}")

    # Clasificación
    _clasificar_df(df_final.copy(), directorio_base, nombre_archivo)

    print(f"\nCompletado en {time.time() - t0:.2f}s — {total_filas} filas")


if __name__ == "__main__":
    print("=== SCRAPER MANIFIESTOS AÉREOS SUNAT ===")
    f_inicio = input("Fecha inicio (dd/mm/yyyy): ").strip()
    f_fin    = input("Fecha fin    (dd/mm/yyyy): ").strip()
    procesar_manifiestos(f_inicio, f_fin)
